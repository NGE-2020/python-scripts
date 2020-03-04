from openpyxl import load_workbook
workbook = load_workbook(filename="nyse.xls")
workbook.sheetnames

sheet = workbook.active
sheet

sheet.title
